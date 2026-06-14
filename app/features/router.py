from __future__ import annotations

import asyncio
import json
import logging
import random
import tempfile
import time
import threading

import httpx
from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import JSONResponse, StreamingResponse, FileResponse
from pathlib import Path
from typing import AsyncIterator

from app.features.prompt import build_extraction_prompt
from app.models.schemas import ExtractResponse, ExtractionMeta, ExtractionResult
from app.services.file_service import decode_txt_bytes, validate_and_read_upload
from app.services.llm_client import LLMClient
from app.services.paddle_ocr import process_pdf
from app.services.reference_service import compare_extraction
from app.core.config import get_settings

router = APIRouter(prefix="/extract", tags=["Extraction"])
llm_client = LLMClient()
settings = get_settings()

logger = logging.getLogger(__name__)

# LLM concurrency — semaphore limit matches pod count.
# Waiting happens here (no timeout clock), not inside the LLM service.
_llm_semaphore: asyncio.Semaphore | None = None


def _get_llm_semaphore() -> asyncio.Semaphore:
    """Lazy-init so the semaphore is created inside the running event loop."""
    global _llm_semaphore
    if _llm_semaphore is None:
        _llm_semaphore = asyncio.Semaphore(settings.llm_max_concurrent)
    return _llm_semaphore

# ---------------------------------------------------------------------------
# Path to txt files from ocr and cleanup function
# ---------------------------------------------------------------------------

_RESULTS_DIR = Path("results").resolve()
_RESULTS_DIR.mkdir(parents=True, exist_ok=True)

def _cleanup_old_ocr_results(max_age_seconds: int = 3600) -> None:
    """Delete OCR text files older than max_age_seconds from results/."""
    now = time.time()
    deleted = 0
    for f in _RESULTS_DIR.glob("paddle_*.txt"):
        try:
            if now - f.stat().st_mtime > max_age_seconds:
                f.unlink()
                deleted += 1
        except OSError:
            pass
    if deleted:
        logger.info("Cleaned up %d old OCR result file(s) from %s", deleted, _RESULTS_DIR)


def _start_cleanup_scheduler(interval_seconds: int = 1800, max_age_seconds: int = 3600) -> None:
    """Run _cleanup_old_ocr_results on a background thread, every interval_seconds."""
    def _loop():
        while True:
            time.sleep(interval_seconds)
            _cleanup_old_ocr_results(max_age_seconds)

    t = threading.Thread(target=_loop, daemon=True, name="ocr-cleanup")
    t.start()
    logger.info("OCR cleanup scheduler started (interval=%ds, max_age=%ds)", interval_seconds, max_age_seconds)


# ---------------------------------------------------------------------------
# PaddleOCR helper — runs OCR in-process via paddle_ocr.process_pdf
# ---------------------------------------------------------------------------

async def _pdf_to_text_via_paddleocr(pdf_bytes: bytes, filename: str, timeout: float | None = None) -> str:
    """
    Write PDF bytes to a temp file, run PaddleOCR in-process, and return
    the extracted plain text.
    """
    import asyncio

    logger.debug("Running in-process PaddleOCR on '%s' (%d bytes)", filename, len(pdf_bytes))

    suffix = Path(filename).suffix or ".pdf"
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    try:
        tmp.write(pdf_bytes)
        tmp.close()

        loop = asyncio.get_running_loop()
        text = await loop.run_in_executor(None, process_pdf, tmp.name)

        logger.debug("PaddleOCR returned %d characters for '%s'", len(text), filename)

        # Save OCR text to results/ so the browser can download it
        stem = Path(filename).stem.lower().replace(" ", "_")
        txt_filename = f"paddle_{stem}.txt"
        out_path = _RESULTS_DIR / txt_filename
        out_path.write_text(text, encoding="utf-8")
        logger.debug("OCR text saved to %s", out_path)

        return text
    finally:
        Path(tmp.name).unlink(missing_ok=True)


# ---------------------------------------------------------------------------
# Core extraction pipeline (called by both single-file and batch routes)
# ---------------------------------------------------------------------------

async def _run_extraction(original_text: str, source: str, timeout: float | None = None) -> ExtractResponse:
    prompt = build_extraction_prompt(original_text)
    llm_result = await llm_client.extract_fields(
        prompt,
        stop=[
            "} {",
            "\n} {",
            "\n}{",
            "}\n{",
            "}\r\n{",
            "}\n\n",
            "}\r\n\r\n",
            "}\n ",
            "} \n",
            "}\n#",
            "}\n`",
            "\n}\n ",
            "\n}\n#",
            "\n}\n`",
            "\n}\n\n",
            "\n}\r\n\r\n",
        ],
        timeout=timeout,
    )

    extracted = ExtractionResult(
        name=llm_result.get("name"),
        master_account_number=llm_result.get("master_account_number"),
        sub_account_number=llm_result.get("sub_account_number"),
        address=llm_result.get("address"),
        fi_num=llm_result.get("fi_num"),
        bank_name=llm_result.get("bank_name"),
    )

    comparison = compare_extraction(
        filename_raw=source,
        bank_name=extracted.bank_name,
        fi_num=extracted.fi_num,
        master_account_number=extracted.master_account_number,
        sub_account_number=extracted.sub_account_number,
    )

    return ExtractResponse(
        success=True,
        message="Extraction completed successfully.",
        data=extracted,
        meta=ExtractionMeta(
            input_characters=len(original_text),
            llm_called=True,
            source=source,
        ),
        comparison=comparison,
    )


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@router.post("/health")
async def health_check() -> dict[str, str]:
    return {"status": "ok"}


@router.post("/from-file", response_model=ExtractResponse)
async def extract_from_file(file: UploadFile = File(...)) -> ExtractResponse:
    raw_bytes, ext = await validate_and_read_upload(file)
    filename = file.filename or "uploaded_file"

    if ext == ".pdf":
        original_text = await _pdf_to_text_via_paddleocr(raw_bytes, filename)
    else:
        original_text = decode_txt_bytes(raw_bytes)

    return await _run_extraction(original_text=original_text, source=filename)

@router.post("/from-text", response_model=ExtractResponse)
async def extract_from_text(
    text: str = Form(...),
    filename: str = Form(...),
) -> ExtractResponse:
    """Accept raw OCR text and filename, run LLM extraction, return result."""
    return await _run_extraction(original_text=text, source=filename)

@router.post("/ocr-only")
async def ocr_only(file: UploadFile = File(...)) -> JSONResponse:
    """
    Run PaddleOCR on a PDF and return the extracted text.
    Does NOT call the LLM. The frontend uses this to decouple OCR from extraction.
    
    Response: { "status": "done"|"error", "text": str, "txt_filename": str, "error": str|null }
    """
    from fastapi.responses import JSONResponse

    raw_bytes, ext = await validate_and_read_upload(file)
    filename = file.filename or "uploaded_file"

    if ext != ".pdf":
        return JSONResponse({"status": "error", "text": None, "txt_filename": None, "error": "Only PDF files are supported."})

    try:
        text = await _pdf_to_text_via_paddleocr(raw_bytes, filename)
        stem = Path(filename).stem.lower().replace(" ", "_")
        txt_filename = f"paddle_{stem}.txt"
        return JSONResponse({"status": "done", "text": text, "txt_filename": txt_filename, "error": None})
    except Exception as exc:
        return JSONResponse({"status": "error", "text": None, "txt_filename": None, "error": str(exc)})
    
@router.get("/ocr-download/{filename}")
async def ocr_download(filename: str) -> FileResponse:
    """
    Download a previously saved OCR text file from results/.
    Prevents path traversal by resolving inside _RESULTS_DIR.
    """
    safe_name = Path(filename).name
    file_path = _RESULTS_DIR / safe_name

    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(status_code=404, detail=f"File '{safe_name}' not found.")

    return FileResponse(
        path=str(file_path),
        media_type="text/plain",
        filename=safe_name,
    )

# ---------------------------------------------------------------------------
# The Protected LLM Worker
# ---------------------------------------------------------------------------

async def _llm_task(
    index: int,
    total: int,
    filename: str,
    ocr_text: str,
    out_queue: asyncio.Queue,
) -> None:
    """
    Acquire semaphore slot → call LLM → retry on failure → push event to queue.

    Waiting on the semaphore is free (no timeout clock).
    The 600s timeout only starts when the request actually enters the LLM service.
    Retries get a fresh 600s clock each time.
    """
    sem       = _get_llm_semaphore()
    last_exc: Exception | None = None

    for attempt in range(settings.llm_max_retries + 1):
        if attempt > 0:
            backoff = settings.llm_retry_base_backoff * (2 ** (attempt - 1))
            jitter  = random.uniform(0.0, 1.0)
            wait    = backoff + jitter
            logger.warning(
                "LLM retry %d/%d for '%s' — waiting %.1fs",
                attempt, settings.llm_max_retries, filename, wait,
            )
            # Slot is NOT held during backoff sleep.
            # Other waiting files can acquire it freely.
            await asyncio.sleep(wait)

        # Re-acquire slot fresh for each attempt.
        # If other files are waiting, they compete fairly here.
        async with sem:
            try:
                result = await _run_extraction(original_text=ocr_text, source=filename)
            except Exception as exc:
                # Slot released immediately on exception (end of async with)
                last_exc = exc
                logger.warning(
                    "LLM attempt %d/%d failed for '%s': %s",
                    attempt + 1, settings.llm_max_retries + 1, filename, exc,
                )
                continue  # go to next attempt, slot is already free

        # Success — slot already released by async with
        await out_queue.put({
            "index": index, "total": total, "filename": filename,
            "stage": "llm_done", "result": result.model_dump(),
        })
        return

    # All retries exhausted
    logger.error(
        "LLM permanently failed for '%s' after %d attempts: %s",
        filename, settings.llm_max_retries + 1, last_exc,
    )
    await out_queue.put({
        "index": index, "total": total, "filename": filename,
        "stage": "error", "error": str(last_exc),
    })

# ---------------------------------------------------------------------------
# Stream Batch Pipeline (OCR, LLM, CSV)
# ---------------------------------------------------------------------------

async def _stream_batch_pipeline(files: list[UploadFile]) -> AsyncIterator[str]:
    """
    OCR runs sequentially — one file at a time.
    LLM tasks fire as each OCR completes and run concurrently,
    capped at llm_max_concurrent (semaphore) so no file ever queues
    inside the LLM service — eliminating timeout risk.
    Results stream to the browser as each task settles.
    """
    total      = len(files)
    llm_tasks: list[asyncio.Task] = []
    out_queue: asyncio.Queue      = asyncio.Queue()
    audit_records: list[dict]     = []

    # --- OCR loop (sequential) ---
    for index, upload in enumerate(files, start=1):
        filename = upload.filename or f"file_{index}"

        try:
            raw_bytes, ext = await validate_and_read_upload(upload)
            if ext != ".pdf":
                await out_queue.put({
                    "index": index, "total": total, "filename": filename,
                    "stage": "error", "error": "Only PDF files are supported.",
                })
                # fire a no-op task so task count stays consistent
                llm_tasks.append(asyncio.create_task(asyncio.sleep(0)))
                continue

            ocr_text = await _pdf_to_text_via_paddleocr(raw_bytes, filename)
            stem         = Path(filename).stem.lower().replace(" ", "_")
            txt_filename = f"paddle_{stem}.txt"

            await out_queue.put({
                "index": index, "total": total, "filename": filename,
                "stage": "ocr_done", "txt_filename": txt_filename,
            })

        except Exception as exc:
            logger.warning("OCR failed [%d/%d] '%s': %s", index, total, filename, exc)
            await out_queue.put({
                "index": index, "total": total, "filename": filename,
                "stage": "error", "error": str(exc),
            })
            llm_tasks.append(asyncio.create_task(asyncio.sleep(0)))
            continue

        # Fire LLM task immediately — it will wait on semaphore internally
        task = asyncio.create_task(
            _llm_task(index, total, filename, ocr_text, out_queue)
        )
        llm_tasks.append(task)

    # --- Drain the queue, yield events as they arrive ---
    # We know exactly how many events to expect:
    #   - 1 ocr_done (or error) per file  → already in queue from loop above
    #   - 1 llm_done (or error) per file  → pushed by tasks as they complete
    expected_llm_events = len(llm_tasks)
    llm_events_received = 0

    while llm_events_received < expected_llm_events:
        event = await out_queue.get()
        stage = event.get("stage")

        if stage in ("llm_done", "error"):
            llm_events_received += 1
            if stage == "llm_done":
                audit_records.append({
                    "filename":      event["filename"],
                    "extractResult": event["result"],
                    "extractError":  None,
                })
            else:
                audit_records.append({
                    "filename":      event["filename"],
                    "extractResult": None,
                    "extractError":  event.get("error"),
                })

        yield json.dumps(event) + "\n"

    # --- Audit log (server-side, after all files done) ---
    if audit_records:
        loop = asyncio.get_running_loop()
        await loop.run_in_executor(None, _write_audit_csv, audit_records)

    yield json.dumps({"stage": "batch_done", "total": total}) + "\n"


@router.post("/batch-pipeline")
async def batch_pipeline(
    files: list[UploadFile] = File(...),
) -> StreamingResponse:
    """
    Server-driven batch pipeline: accepts all files, runs OCR→LLM per file,
    streams newline-delimited JSON progress events, writes audit log internally.
    """
    if not files:
        raise HTTPException(status_code=422, detail="No files provided.")

    max_files = settings.max_files_per_batch
    if len(files) > max_files:
        raise HTTPException(
            status_code=422,
            detail=f"Too many files. Received {len(files)}, maximum is {max_files}.",
        )

    logger.info("Batch pipeline request — %d file(s)", len(files))

    return StreamingResponse(
        _stream_batch_pipeline(files),
        media_type="application/x-ndjson",
        headers={"Cache-Control": "no-cache"},
    )


# ---------------------------------------------------------------------------
# Audit log — developer-only, never served back to the browser
# ---------------------------------------------------------------------------

_AUDIT_DIR = Path("audit_logs").resolve()
_AUDIT_DIR.mkdir(parents=True, exist_ok=True)

_AUDIT_FIELDS = [
    ("bank_name",             "Bank Name"),
    ("fi_num",                "FI Code"),
    ("master_account_number", "Master Account No."),
    ("sub_account_number",    "Sub Account No."),
]


def _write_audit_csv(records: list[dict]) -> None:
    """Write one audit XLSX per batch run to audit_logs/ with colour-coded expected columns."""
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = _AUDIT_DIR / f"audit_{timestamp}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit"

    # --- Styles ---
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    green_font = Font(color="006100")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    red_font = Font(color="9C0006")
    total_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # --- Header row ---
    header = ["File Name"]
    for _key, label in _AUDIT_FIELDS:
        header += [f"Extracted {label}", f"Expected {label}"]
    header.append("Accurate Count")

    ws.append(header)
    for col_idx, _ in enumerate(header, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # --- Data rows ---
    total_accurate = 0
    total_fields = 0

    # Track which columns are "Expected" so we can colour them
    # Expected columns are at indices: 3, 5, 7, 9 (1-based) i.e. every even field col
    expected_col_indices = [2 + i * 2 + 1 for i in range(len(_AUDIT_FIELDS))]  # 1-based: 3,5,7,9

    for record in records:
        filename    = record.get("filename", "")
        extract_res = record.get("extractResult") or {}
        extract_err = record.get("extractError")

        if extract_err or not extract_res:
            row_data = [filename] + ["ERROR", ""] * len(_AUDIT_FIELDS) + [0]
            ws.append(row_data)
            row_num = ws.max_row
            for col_idx in range(1, len(row_data) + 1):
                ws.cell(row=row_num, column=col_idx).border = thin_border
            for col_idx in expected_col_indices:
                cell = ws.cell(row=row_num, column=col_idx)
                cell.fill = red_fill
                cell.font = red_font
            total_fields += len(_AUDIT_FIELDS)
            continue

        data = extract_res.get("data") or {}
        cmp  = extract_res.get("comparison") or {}

        row_data = [filename]
        matches = []
        accurate_count = 0
        for key, _label in _AUDIT_FIELDS:
            extracted = (data.get(key) or "").strip()
            expected  = ((cmp.get(key) or {}).get("expected") or "").strip()
            row_data += [extracted, expected]
            is_match = bool(extracted and expected and extracted.upper() == expected.upper())
            matches.append(is_match)
            if is_match:
                accurate_count += 1

        row_data.append(accurate_count)
        total_accurate += accurate_count
        total_fields += len(_AUDIT_FIELDS)
        ws.append(row_data)

        row_num = ws.max_row
        for col_idx in range(1, len(row_data) + 1):
            ws.cell(row=row_num, column=col_idx).border = thin_border

        # Colour expected columns green/red
        for i, col_idx in enumerate(expected_col_indices):
            cell = ws.cell(row=row_num, column=col_idx)
            if matches[i]:
                cell.fill = green_fill
                cell.font = green_font
            else:
                cell.fill = red_fill
                cell.font = red_font

    # --- Summary row ---
    if total_fields > 0:
        pct = (total_accurate / total_fields) * 100
    else:
        pct = 0.0

    summary_row = ["TOTAL"] + [""] * (len(_AUDIT_FIELDS) * 2)
    summary_row.append(f"{total_accurate}/{total_fields} ({pct:.2f}%)")
    ws.append(summary_row)
    row_num = ws.max_row
    for col_idx in range(1, len(summary_row) + 1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.font = total_font
        cell.border = thin_border

    # --- Auto-fit column widths ---
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    wb.save(str(path))
    logger.info("Audit log saved: %s (%d record(s), accuracy %.2f%%)", path, len(records), pct if total_fields else 0.0)